import pathlib
from spreadsheet_extractor import settings
import openpyxl
import pandas as pd

def get_spreadsheet_file(
        data_path,
        file_path,
        read_only
    ):

    data_path = data_path or settings.get_data_path()
    excel_file_path = pathlib.Path(
        pathlib.Path(data_path, settings.EXTRACTS_FOLDER), 
        file_path
    )

    if file_path.endswith('.xlsx'):
        return get_excel_file(excel_file_path, read_only)
    else:
        raise NotImplementedError(f'The extension of the following file is not supported : {file_path}')


def get_excel_file(
    excel_file_path: pathlib.Path,
    read_only: bool
) -> openpyxl.workbook.workbook.Workbook:
    return openpyxl.load_workbook(excel_file_path, data_only=True, read_only=read_only)


def extract_values_from_row(row):
    # return [
    #     c.value.strftime(r"%Y/%m/%d") if c.is_date else c.value \
    #     for c in row
    # ]
    return [
        c.value for c in row
    ]

def get_ls_columns(ws, columns_cells_range):
    return [c.value for c in ws[columns_cells_range][0]]

def get_df_from_cells_range(ws, cell_ranges, ls_columns ):
    dc_ranges = {}
    for cell_range in cell_ranges.split(','):
        dc_ranges[cell_range] = pd.DataFrame(
            list(map(extract_values_from_row, ws[cell_range])),
            columns=ls_columns
        )
    return pd.concat(
        dc_ranges.values(),
        ignore_index=True
    )

def apply_cell_replacement(ws, replace_mapping):
    for cell_range, value in replace_mapping.items():
        ws[cell_range].value = value 
    return ws
